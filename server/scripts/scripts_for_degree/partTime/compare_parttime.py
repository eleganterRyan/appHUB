import pandas as pd
import sys
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

if len(sys.argv) < 4:
    print("用法: python compare_parttime.py <current_file> <last_file> <output_dir>")
    sys.exit(1)

current_file = sys.argv[1]
last_file = sys.argv[2]
output_dir = sys.argv[3]
os.makedirs(output_dir, exist_ok=True)

# 读取数据，移动电话为字符串
cur = pd.read_excel(current_file, dtype={'移动电话': str})
last = pd.read_excel(last_file, dtype={'移动电话': str})

# 检查重复
repeat_cur = cur[cur.duplicated('移动电话', keep=False)]
repeat_last = last[last.duplicated('移动电话', keep=False)]
if not repeat_cur.empty or not repeat_last.empty:
    repeat_info = {
        "cur_repeat": repeat_cur.to_dict(orient='records'),
        "last_repeat": repeat_last.to_dict(orient='records')
    }
    print(json.dumps({"message": "存在重复移动电话", "repeatInfo": repeat_info}, ensure_ascii=False))
    sys.exit(0)

# 检查必需字段
required_cols = ["一级学科名称1", "一级学科名称2", "专业学位类别名称"]
for col in required_cols:
    if col not in cur.columns or col not in last.columns:
        print(json.dumps({"message": f"缺少字段: {col}，请检查上传的Excel文件"}))
        sys.exit(0)

# 主键
key = "移动电话"
cur[key] = cur[key].astype(str)
last[key] = last[key].astype(str)
cur_keys = set(cur[key])
last_keys = set(last[key])

# 1. 在库申请
in_db_keys = cur_keys & last_keys
in_db_rows = []
for k in in_db_keys:
    cur_row = cur[cur[key] == k].iloc[0]
    last_row = last[last[key] == k].iloc[0]
    # 超龄判定
    is_overage = str(cur_row.get("是否超龄", "")).strip() != "否"
    # 专业调整判定
    cur_majors = set([
        str(cur_row.get("一级学科名称1", "")),
        str(cur_row.get("一级学科名称2", "")),
        str(cur_row.get("专业学位类别名称", ""))
    ])
    last_majors = set([
        str(last_row.get("一级学科名称1", "")),
        str(last_row.get("一级学科名称2", "")),
        str(last_row.get("专业学位类别名称", ""))
    ])
    is_major_changed = cur_majors != last_majors
    major_change_str = ";".join(sorted(last_majors)) if is_major_changed else ""
    row = cur_row.to_dict()
    # 插入是否专业调整列
    idx = list(cur.columns).index("是否超龄") + 1 if "是否超龄" in cur.columns else len(cur.columns)
    items = list(row.items())
    items.insert(idx, ("是否专业调整", ("是(" + major_change_str + ")") if is_major_changed else "否"))
    row = dict(items)
    row["_is_overage"] = is_overage
    row["_is_major_changed"] = is_major_changed
    in_db_rows.append(row)

# 排序
in_db_rows_sorted = (
    [r for r in in_db_rows if r["_is_overage"]] +
    [r for r in in_db_rows if not r["_is_overage"] and r["_is_major_changed"]] +
    [r for r in in_db_rows if not r["_is_overage"] and not r["_is_major_changed"]]
)

# 2. 本年度新增
cur_only_keys = cur_keys - last_keys
cur_only_rows = []
for k in cur_only_keys:
    cur_row = cur[cur[key] == k].iloc[0]
    is_overage = str(cur_row.get("是否超龄", "")).strip() != "否"
    row = cur_row.to_dict()
    row["_is_overage"] = is_overage
    cur_only_rows.append(row)
cur_only_rows_sorted = [r for r in cur_only_rows if r["_is_overage"]] + [r for r in cur_only_rows if not r["_is_overage"]]

# 3. 本年度退出
exit_keys = last_keys - cur_keys
exit_rows = []
for k in exit_keys:
    last_row = last[last[key] == k].iloc[0]
    exit_rows.append(last_row)

# 输出Excel
in_db_df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in in_db_rows_sorted])
cur_only_df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in cur_only_rows_sorted])
exit_df = pd.DataFrame(exit_rows)

in_db_path = os.path.join(output_dir, "in_db.xlsx")
cur_only_path = os.path.join(output_dir, "current_only.xlsx")
exit_path = os.path.join(output_dir, "exit.xlsx")

in_db_df.to_excel(in_db_path, index=False)
cur_only_df.to_excel(cur_only_path, index=False)
exit_df.to_excel(exit_path, index=False)

# 标红样式
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
def mark_red_column(path, cond_func, col_name):
    wb = load_workbook(path)
    ws = wb.active
    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = idx
            break
    if col_idx is None:
        wb.save(path)
        return
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        if cond_func(i-2):
            cell = row[col_idx-1]
            cell.fill = red_fill
    wb.save(path)

# 标红：在库申请
mark_red_column(in_db_path, lambda idx: in_db_rows_sorted[idx]["_is_overage"], "是否超龄")
mark_red_column(in_db_path, lambda idx: in_db_rows_sorted[idx]["_is_major_changed"], "是否专业调整")
# 标红：本年度新增
mark_red_column(cur_only_path, lambda idx: cur_only_rows_sorted[idx]["_is_overage"], "是否超龄")

# 统计信息
stats = {
    "in_db_count": len(in_db_rows_sorted),
    "in_db_overage": sum(r["_is_overage"] for r in in_db_rows_sorted),
    "in_db_major_changed": sum(r["_is_major_changed"] for r in in_db_rows_sorted),
    "cur_only_count": len(cur_only_rows_sorted),
    "cur_only_overage": sum(r["_is_overage"] for r in cur_only_rows_sorted),
    "exit_count": len(exit_rows)
}

print(json.dumps({"status": "ok", "stats": stats})) 