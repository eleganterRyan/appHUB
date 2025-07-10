import pandas as pd
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 用法: python compare.py <current_file> <last_file> <output_dir>
if len(sys.argv) < 4:
    print("用法: python compare.py <current_file> <last_file> <output_dir>")
    sys.exit(1)

current_file = sys.argv[1]
last_file = sys.argv[2]
output_dir = sys.argv[3]

os.makedirs(output_dir, exist_ok=True)

# 读取数据，强制工号为字符串
cur = pd.read_excel(current_file, dtype={'工号': str})
last = pd.read_excel(last_file, dtype={'工号': str})

# 主键
key = "工号"

# 再次确保工号为字符串
cur[key] = cur[key].astype(str)
last[key] = last[key].astype(str)

# 专业字段
major_cols = ["招生专业1", "招生专业2", "招生专业3"]

# 交集、差集
cur_keys = set(cur[key])
last_keys = set(last[key])

in_db_keys = cur_keys & last_keys
cur_only_keys = cur_keys - last_keys
exit_keys = last_keys - cur_keys

# 辅助函数：获取专业集合
def get_major_set(row):
    return set(str(row[col]).strip() for col in major_cols if pd.notna(row[col]) and str(row[col]).strip())

# 1. 在库申请
in_db_rows = []
for k in in_db_keys:
    cur_row = cur[cur[key] == k].iloc[0]
    last_row = last[last[key] == k].iloc[0]
    # 超龄判定
    is_overage = str(cur_row.get("是否超龄", "")).strip() != "否"
    # 专业调整判定
    cur_majors = get_major_set(cur_row)
    last_majors = get_major_set(last_row)
    is_major_changed = cur_majors != last_majors
    major_change_str = ";".join(sorted(last_majors)) if is_major_changed else ""
    in_db_rows.append({
        **cur_row,
        "是否超龄": cur_row.get("是否超龄", ""),
        "是否专业调整": ("是(" + major_change_str + ")") if is_major_changed else "否",
        "_is_overage": is_overage,
        "_is_major_changed": is_major_changed
    })

# 排序：超龄→专业调整→其他
in_db_rows_sorted = (
    [r for r in in_db_rows if r["_is_overage"]] +
    [r for r in in_db_rows if not r["_is_overage"] and r["_is_major_changed"]] +
    [r for r in in_db_rows if not r["_is_overage"] and not r["_is_major_changed"]]
)

# 2. 本年度申请
cur_only_rows = []
for k in cur_only_keys:
    cur_row = cur[cur[key] == k].iloc[0]
    is_overage = str(cur_row.get("是否超龄", "")).strip() != "否"
    cur_only_rows.append({
        **cur_row,
        "是否超龄": cur_row.get("是否超龄", ""),
        "_is_overage": is_overage
    })
cur_only_rows_sorted = (
    [r for r in cur_only_rows if r["_is_overage"]] +
    [r for r in cur_only_rows if not r["_is_overage"]]
)

# 3. 本年度退出
exit_rows = []
for k in exit_keys:
    last_row = last[last[key] == k].iloc[0]
    exit_rows.append(last_row)

# 输出Excel
in_db_df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in in_db_rows_sorted])
cur_only_df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in cur_only_rows_sorted])
exit_df = pd.DataFrame(exit_rows)

in_db_path = os.path.join(output_dir, "in_db.xlsx")
cur_only_path = os.path.join(output_dir, "current_only.xlsx")
exit_path = os.path.join(output_dir, "exit.xlsx")

in_db_df.to_excel(in_db_path, index=False)
cur_only_df.to_excel(cur_only_path, index=False)
exit_df.to_excel(exit_path, index=False)

# 标红样式
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

def mark_red_column(path, cond_func, col_name):
    wb = load_workbook(path)
    ws = wb.active
    # 找到目标列索引
    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            col_idx = idx
            break
    if col_idx is None:
        wb.save(path)
        return
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        if cond_func(i-2):
            cell = row[col_idx-1]
            cell.fill = red_fill
    wb.save(path)

def mark_red(path, cond_func):
    wb = load_workbook(path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        if cond_func(i-2):
            for cell in row:
                cell.fill = red_fill
    wb.save(path)

# 标红：在库申请（超龄只标“是否超龄”列，专业调整只标“是否专业调整”列）
mark_red_column(in_db_path, lambda idx: in_db_rows_sorted[idx]["_is_overage"], "是否超龄")
mark_red_column(in_db_path, lambda idx: in_db_rows_sorted[idx]["_is_major_changed"], "是否专业调整")
# 标红：本年度申请（超龄只标“是否超龄”列）
mark_red_column(cur_only_path, lambda idx: cur_only_rows_sorted[idx]["_is_overage"], "是否超龄")

print("对比完成", in_db_path, cur_only_path, exit_path) 