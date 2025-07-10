import sys
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "缺少参数"}))
        sys.exit(1)
    try:
        args = json.loads(sys.argv[1])
        files = args.get("files", [])
        output = args.get("output")
        if not files or not output:
            print(json.dumps({"success": False, "error": "参数不完整"}))
            sys.exit(1)
        wb_out = Workbook()
        wb_out.remove(wb_out.active)  # 移除默认sheet
        sheet_names = set()
        for file_path in files:
            try:
                wb = load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 保证sheet名唯一
                    base_name = sheet_name
                    new_name = base_name
                    idx = 1
                    while new_name in sheet_names:
                        new_name = f"{base_name}_{idx}"
                        idx += 1
                    sheet_names.add(new_name)
                    ws_out = wb_out.create_sheet(new_name)
                    for row in ws.iter_rows(values_only=True):
                        ws_out.append(row)
            except Exception as e:
                continue  # 跳过有问题的文件
        if not wb_out.sheetnames:
            print(json.dumps({"success": False, "error": "没有有效sheet"}))
            sys.exit(1)
        os.makedirs(os.path.dirname(output), exist_ok=True)
        wb_out.save(output)
        print(json.dumps({"success": True}))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main() 